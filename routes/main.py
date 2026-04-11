from flask import Blueprint, make_response, redirect, render_template,url_for,flash, request ,jsonify , current_app, send_from_directory
from weasyprint import HTML
from flask_login import login_required, current_user
from openpyxl import load_workbook
import io
import os
from uuid import uuid4
from werkzeug.utils import secure_filename
from services.storage_service import subir_pdf


from services.catalog_service import (
    agregar_asunto,
    agregar_empresa,
    agregar_ingeniero,
    eliminar_asunto,
    eliminar_empresa,
    eliminar_ingeniero,
    listar_catalogos,
    obtener_asuntos_sugeridos,
    obtener_empresas,
    obtener_ingenieros_por_empresa,
    actualizar_empresa,
)
from services.cotizacion_service import (
    actualizar_cotizacion,
    actualizar_estado,
    construir_params_duplicado,
    construir_params_edicion,
    crear_cotizacion,
    eliminar_cotizacion,
    incrementar_correlativo,
    listar_cotizaciones,
    obtener_cotizacion,
    obtener_no_referencia,
    preparar_data_pdf,
)

from services.orden_service import (
    obtener_ordenes,
    obtener_empresas_activas,
    crear_orden_compra,
    obtener_orden_por_id,
    registrar_factura_orden,
    marcar_factura_pagada,
    guardar_resultado_extraccion_orden,
    aplicar_datos_parseados_a_orden,
    obtener_empresa_por_id,
    generar_expediente_unificado
    )

from services.helpers import fecha_guatemala, fecha_validez
from services.pdf_ocr_service import extraer_texto_pdf
from services.ocr_parsers import parsear_orden_por_plantilla
from services.factura_pdf_parser import parsear_factura_sat


bp = Blueprint("main", __name__)

@bp.before_request
def require_login():
    from flask_login import current_user
    if not current_user.is_authenticated:
        from flask import redirect, url_for
        return redirect(url_for("auth.login"))


def _es_url_remota(valor: str) -> bool:
    return isinstance(valor, str) and (
        valor.startswith("http://") or valor.startswith("https://")
    )


def _resolver_ruta_pdf(valor: str, upload_folder: str) -> str:
    if not valor:
        return ""
    valor = valor.strip()
    if _es_url_remota(valor):
        return valor
    if os.path.isabs(valor):
        return valor
    return os.path.join(upload_folder, valor)


def _resolver_url_publica_archivo(valor: str, endpoint: str) -> str:
    if not valor:
        return ""
    valor = valor.strip()
    if _es_url_remota(valor):
        return valor
    return url_for(endpoint, nombre_archivo=os.path.basename(valor))

@bp.route("/")
def index():
    fecha = fecha_guatemala()
    validez = fecha_validez()
    no_referencia = obtener_no_referencia()

    datos_fijos = {
        "precio_texto": "Se entienden netos en Quetzales",
        "tiempo_entrega": "Inmediata",
        "encargado": "Melvin Siney",
        "contacto_nombre": "Melvin Siney",
        "contacto_telefono": "5502-5762",
        "contacto_correo": "melvinsiney@gmail.com",
    }

    duplicado = {
        "id": request.args.get("id", ""),
        "modo_edicion": request.args.get("modo_edicion", "0"),
        "empresa": request.args.get("empresa", ""),
        "ingeniero": request.args.get("ingeniero", ""),
        "su_referencia": request.args.get("su_referencia", ""),
        "asunto": request.args.get("asunto", ""),
        "encargado": request.args.get("encargado", datos_fijos["encargado"]),
        "contacto_nombre": request.args.get("contacto_nombre", datos_fijos["contacto_nombre"]),
        "contacto_telefono": request.args.get("contacto_telefono", datos_fijos["contacto_telefono"]),
        "contacto_correo": request.args.get("contacto_correo", datos_fijos["contacto_correo"]),
        "items_json": request.args.get("items_json", "[]"),
        "precio_texto": request.args.get("precio_texto", datos_fijos["precio_texto"]),
        "tiempo_entrega": request.args.get("tiempo_entrega", datos_fijos["tiempo_entrega"]),
        "validez": request.args.get("validez", validez),
        "no_referencia": request.args.get("no_referencia", no_referencia),
        "fecha": request.args.get("fecha", fecha),
    }

    return render_template(
        "index.html",
        fecha=duplicado["fecha"],
        validez=duplicado["validez"],
        no_referencia=duplicado["no_referencia"],
        datos_fijos=datos_fijos,
        duplicado=duplicado,
        empresas=obtener_empresas(),
    )


@bp.route("/exportar-pdf", methods=["POST"])
def exportar_pdf():
    data = request.form.to_dict()
    crear_cotizacion(data)
    html = render_template("pdf_template.html", data=data)
    pdf = HTML(string=html).write_pdf()
    incrementar_correlativo()

    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"inline; filename={data.get('no_referencia', 'cotizacion')}.pdf"
    return response


@bp.route("/cotizaciones")
def cotizaciones():
    cotizaciones_data, total_mes = listar_cotizaciones()
    return render_template("cotizaciones.html", cotizaciones=cotizaciones_data, total_mes=total_mes)


@bp.route("/cambiar-estado/<int:id>", methods=["POST"])
def cambiar_estado(id):
    actualizar_estado(id, request.form.get("estado", "Pendiente"))
    return redirect("/cotizaciones")


@bp.route("/descargar-pdf/<int:id>")
def descargar_pdf_historial(id):
    cotizacion = obtener_cotizacion(id)
    if not cotizacion:
        return "Cotización no encontrada", 404

    html = render_template("pdf_template.html", data=preparar_data_pdf(cotizacion))
    pdf = HTML(string=html).write_pdf()

    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"inline; filename={cotizacion['no_referencia']}.pdf"
    return response


@bp.route("/duplicar/<int:id>")
def duplicar_cotizacion(id):
    cotizacion = obtener_cotizacion(id)
    if not cotizacion:
        return "Cotización no encontrada", 404
    return redirect(f"/?{construir_params_duplicado(cotizacion)}")


@bp.route("/editar/<int:id>")
def editar_cotizacion(id):
    cotizacion = obtener_cotizacion(id)
    if not cotizacion:
        return "Cotización no encontrada", 404
    return redirect(f"/?{construir_params_edicion(cotizacion)}")


@bp.route("/guardar-edicion", methods=["POST"])
def guardar_edicion():
    data = request.form.to_dict()
    if not data.get("cotizacion_id"):
        return "ID de cotización no recibido", 400
    actualizar_cotizacion(data)
    return redirect("/cotizaciones")


@bp.route("/eliminar/<int:id>", methods=["POST"])
def eliminar(id):
    eliminar_cotizacion(id)
    return redirect("/cotizaciones")


@bp.route("/api/ingenieros/<int:empresa_id>")
def api_ingenieros(empresa_id):
    ingenieros = obtener_ingenieros_por_empresa(empresa_id)
    return {
        "ingenieros": [
            {
                "id": i["id"],
                "nombre": i["nombre"],
                "titulo": i["titulo"] or "",
                "nombre_mostrar": f"{(i['titulo'] or '').strip()} {(i['nombre'] or '').strip()}".strip(),
            }
            for i in ingenieros
        ]
    }


@bp.route("/api/asuntos")
def api_asuntos():
    empresa_id = request.args.get("empresa_id", type=int)
    asuntos = obtener_asuntos_sugeridos(empresa_id)
    return {"asuntos": [a["asunto"] for a in asuntos]}

@bp.route("/ordenes")
def ordenes():
    ordenes_data = obtener_ordenes()
    return render_template("ordenes.html", ordenes=ordenes_data)


@bp.route("/catalogos")
def catalogos():
    empresas, ingenieros, asuntos = listar_catalogos()
    return render_template("catalogos.html", empresas=empresas, ingenieros=ingenieros, asuntos=asuntos)


@bp.route("/agregar-empresa", methods=["POST"])
def crear_empresa():
    nombre = request.form.get("nombre", "").strip()
    nit = request.form.get("nit", "").strip()
    direccion = request.form.get("direccion", "").strip()
    dias_credito = request.form.get("dias_credito", "30").strip()
    plantilla_ocr = ""
    formato_descripcion_sat ="por servicios tecnicos segun orden de compra {numero_orden} {proyecto} "
    activo = request.form.get("activo") == "on"

    if nombre:
        agregar_empresa(
            nombre=nombre,
            nit=nit,
            direccion=direccion,
            dias_credito=dias_credito,
            plantilla_ocr=plantilla_ocr,
            formato_descripcion_sat=formato_descripcion_sat,
            activo=activo,
        )

    return redirect("/catalogos")


@bp.route("/agregar-ingeniero", methods=["POST"])
def crear_ingeniero():
    empresa_id = request.form.get("empresa_id", "").strip()
    titulo = request.form.get("titulo", "").strip()
    nombre = request.form.get("nombre", "").strip()
    if empresa_id and nombre:
        agregar_ingeniero(int(empresa_id), titulo, nombre)
    return redirect("/catalogos")


@bp.route("/agregar-asunto", methods=["POST"])
def crear_asunto():
    empresa_id = request.form.get("empresa_id", "").strip()
    asunto = request.form.get("asunto", "").strip()
    if asunto:
        agregar_asunto(None if empresa_id == "" else int(empresa_id), asunto)
    return redirect("/catalogos")


@bp.route("/eliminar-empresa/<int:id>", methods=["POST"])
def borrar_empresa(id):
    eliminar_empresa(id)
    return redirect("/catalogos")


@bp.route("/eliminar-ingeniero/<int:id>", methods=["POST"])
def borrar_ingeniero(id):
    eliminar_ingeniero(id)
    return redirect("/catalogos")


@bp.route("/eliminar-asunto/<int:id>", methods=["POST"])
def borrar_asunto(id):
    eliminar_asunto(id)
    return redirect("/catalogos")

@bp.route("/editar-empresa/<int:id>", methods=["POST"])
def editar_empresa(id):
    nombre = request.form.get("nombre", "").strip()
    nit = request.form.get("nit", "").strip()
    direccion = request.form.get("direccion", "").strip()
    dias_credito = request.form.get("dias_credito", "30").strip()
    plantilla_ocr = request.form.get("plantilla_ocr", "").strip()
    formato_descripcion_sat = request.form.get("formato_descripcion_sat", "").strip()
    activo = request.form.get("activo") == "on"

    if nombre:
        actualizar_empresa(
            empresa_id=id,
            nombre=nombre,
            nit=nit,
            direccion=direccion,
            dias_credito=dias_credito,
            plantilla_ocr=plantilla_ocr,
            formato_descripcion_sat=formato_descripcion_sat,
            activo=activo,
        )

    return redirect("/catalogos")

@bp.route("/importar-excel", methods=["POST"])
def importar_excel():
    archivo = request.files.get("archivo")

    if not archivo:
        return jsonify({"ok": False, "error": "No se envió ningún archivo."}), 400

    try:
        contenido = io.BytesIO(archivo.read())
        wb = load_workbook(contenido, data_only=True)
        ws = wb.active

        filas = list(ws.iter_rows(values_only=True))
        if not filas or len(filas) < 2:
            return jsonify({"ok": False, "error": "El archivo está vacío o no tiene datos."}), 400

        encabezados = [str(c).strip().lower() if c is not None else "" for c in filas[0]]

        try:
            idx_cantidad = encabezados.index("cantidad")
            idx_descripcion = encabezados.index("descripcion")
            idx_precio = encabezados.index("precio_unitario")
        except ValueError:
            return jsonify({
                "ok": False,
                "error": "El Excel debe tener las columnas: cantidad, descripcion, precio_unitario"
            }), 400

        items = []
        for fila in filas[1:]:
            if not fila:
                continue

            cantidad = fila[idx_cantidad] if idx_cantidad < len(fila) else None
            descripcion = fila[idx_descripcion] if idx_descripcion < len(fila) else None
            precio = fila[idx_precio] if idx_precio < len(fila) else None

            if descripcion in (None, ""):
                continue

            try:
                cantidad = float(cantidad or 0)
            except Exception:
                cantidad = 0

            try:
                precio = float(precio or 0)
            except Exception:
                precio = 0

            items.append({
                "cantidad": cantidad,
                "descripcion": str(descripcion).strip(),
                "precio_unitario": precio
            })

        return jsonify({"ok": True, "items": items})

    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al leer el Excel: {str(e)}"}), 500
    
@bp.route("/ordenes/nueva", methods=["GET", "POST"])
def nueva_orden():
    if request.method == "POST":
        try:
            archivo_orden = request.form.get("archivo_orden_guardado", "").strip()
            archivo = request.files.get("archivo_orden")

            if archivo and archivo.filename:
                if not archivo_permitido_pdf(archivo.filename):
                    flash("Solo se permiten archivos PDF para la orden de compra.", "warning")
                    return redirect("/ordenes/nueva")

                archivo_orden = guardar_pdf_temporal_orden(archivo)

            data = {
                "empresa_id": int(request.form.get("empresa_id")),
                "numero_orden": request.form.get("numero_orden", "").strip(),
                "fecha_orden": request.form.get("fecha_orden", "").strip(),
                "subtotal": request.form.get("subtotal", "0").strip(),
                "iva": request.form.get("iva", "0").strip(),
                "total": request.form.get("total", "0").strip(),
                "moneda": request.form.get("moneda", "GTQ").strip(),
                "proyecto": request.form.get("proyecto", "").strip(),
                "descripcion_extraida": request.form.get("descripcion_extraida", "").strip(),
                "condiciones_pago": request.form.get("condiciones_pago", "").strip(),
                "dias_credito": request.form.get("dias_credito", "").strip(),
                "archivo_orden": archivo_orden,
            }

            if not data["empresa_id"]:
                flash("Debes seleccionar una empresa.", "warning")
                return redirect("/ordenes/nueva")

            if not data["numero_orden"]:
                flash("Debes ingresar el número de orden.", "warning")
                return redirect("/ordenes/nueva")

            crear_orden_compra(data)
            flash("Orden de compra creada correctamente.", "success")
            return redirect("/ordenes")

        except Exception as e:
            flash(f"Error al crear la orden: {e}", "danger")
            return redirect("/ordenes/nueva")

    empresas = obtener_empresas_activas()
    return render_template(
        "orden_nueva.html",
        empresas=empresas,
        form_data={},
        archivo_orden_guardado="",
        datos_detectados={},
    )

@bp.route("/ordenes/<int:orden_id>")
def detalle_orden(orden_id):
    orden = obtener_orden_por_id(orden_id)
    if not orden:
        flash("La orden no existe.", "warning")
        return redirect("/ordenes")

    return render_template(
        "orden_detalle.html",
        orden=orden,
        archivo_orden_url=_resolver_url_publica_archivo(
            orden.get("archivo_orden"),
            "main.ver_archivo_orden",
        ),
        archivo_factura_url=_resolver_url_publica_archivo(
            orden.get("archivo_factura"),
            "main.ver_archivo_factura",
        ),
        archivo_expediente_url=_resolver_url_publica_archivo(
            orden.get("archivo_expediente"),
            "main.ver_archivo_factura",
        ),
    )

@bp.route("/ordenes/<int:orden_id>/factura", methods=["GET", "POST"])
def registrar_factura(orden_id):
    orden = obtener_orden_por_id(orden_id)
    if not orden:
        flash("La orden no existe.", "warning")
        return redirect("/ordenes")

    if request.method == "POST":
        try:
            archivo_factura_guardado = request.form.get("archivo_factura_guardado", "").strip()
            archivo = request.files.get("archivo_factura")

            if archivo and archivo.filename:
                archivo_factura_guardado = guardar_pdf_temporal_factura(archivo)
                if not archivo_factura_guardado:
                    flash("Solo se permiten archivos PDF para la factura.", "warning")
                    return redirect(f"/ordenes/{orden_id}/factura")

            data = {
                "numero_factura": request.form.get("numero_factura", "").strip(),
                "fecha_factura": request.form.get("fecha_factura", "").strip(),
                "nit_facturado": request.form.get("nit_facturado", "").strip(),
                "direccion_facturada": request.form.get("direccion_facturada", "").strip(),
                "descripcion_sat": request.form.get("descripcion_sat", "").strip(),
                "monto_facturado": request.form.get("monto_facturado", "").strip(),
                "observaciones": request.form.get("observaciones", "").strip(),
                "archivo_factura": archivo_factura_guardado,
                "archivo_expediente": "",
            }

            if archivo_factura_guardado and orden.get("archivo_orden"):
                nombre_expediente = f"expediente_orden_{orden_id}_{uuid4().hex}.pdf"
                ruta_expediente = os.path.join(current_app.config["UPLOAD_FOLDER_FACTURAS"], nombre_expediente)

                ruta_orden = _resolver_ruta_pdf(
                    orden["archivo_orden"],
                    current_app.config["UPLOAD_FOLDER_ORDENES"],
                )
                ruta_factura = _resolver_ruta_pdf(
                    archivo_factura_guardado,
                    current_app.config["UPLOAD_FOLDER_FACTURAS"],
                )

                generar_expediente_unificado(ruta_orden, ruta_factura, ruta_expediente)
                try:
                    url_expediente = subir_pdf(ruta_expediente, carpeta="cotizador/expedientes")
                    os.remove(ruta_expediente)
                    data["archivo_expediente"] = url_expediente
                except Exception:
                    data["archivo_expediente"] = nombre_expediente

            registrar_factura_orden(orden_id, data)
            flash("Factura registrada correctamente.", "success")
            return redirect(f"/ordenes/{orden_id}")

        except Exception as e:
            flash(f"Error al registrar la factura: {e}", "danger")
            return redirect(f"/ordenes/{orden_id}/factura")

    form_data = {
        "numero_factura": orden.get("numero_factura") or "",
        "fecha_factura": orden.get("fecha_factura") or "",
        "monto_facturado": orden.get("monto_facturado") if orden.get("monto_facturado") is not None else (orden.get("total") or ""),
        "nit_facturado": orden.get("nit_facturado") or orden.get("empresa_nit") or "",
        "direccion_facturada": orden.get("direccion_facturada") or orden.get("empresa_direccion") or "",
        "descripcion_sat": orden.get("descripcion_sat") or orden.get("descripcion_sat_sugerida") or "",
        "observaciones": orden.get("observaciones") or "",
    }

    return render_template(
        "registrar_factura.html",
        orden=orden,
        form_data=form_data,
        datos_detectados={},
        archivo_factura_guardado=orden.get("archivo_factura") or "",
    )

@bp.route("/ordenes/<int:orden_id>/pagar", methods=["GET", "POST"])
def pagar_orden(orden_id):
    orden = obtener_orden_por_id(orden_id)
    if not orden:
        flash("La orden no existe.", "warning")
        return redirect("/ordenes")

    if not orden.get("numero_factura"):
        flash("Primero debes registrar la factura.", "warning")
        return redirect(f"/ordenes/{orden_id}")

    if request.method == "POST":
        try:
            fecha_pago_real = request.form.get("fecha_pago_real", "").strip()
            observaciones = request.form.get("observaciones", "").strip()

            if not fecha_pago_real:
                flash("Debes ingresar la fecha real de pago.", "warning")
                return redirect(f"/ordenes/{orden_id}/pagar")

            marcar_factura_pagada(orden_id, fecha_pago_real, observaciones)
            flash("Pago registrado correctamente.", "success")
            return redirect(f"/ordenes/{orden_id}")

        except Exception as e:
            flash(f"Error al registrar el pago: {e}", "danger")
            return redirect(f"/ordenes/{orden_id}/pagar")

    return render_template("marcar_pagada.html", orden=orden)

def archivo_permitido_pdf(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() == "pdf"

def guardar_pdf_temporal_factura(archivo):
    if not archivo or not archivo.filename:
        return ""
    if not archivo_permitido_pdf(archivo.filename):
        return ""
    nombre_seguro = secure_filename(archivo.filename)
    nombre_final = f"{uuid4().hex}_{nombre_seguro}"
    ruta_temp = os.path.join(current_app.config["UPLOAD_FOLDER_FACTURAS"], nombre_final)
    archivo.save(ruta_temp)
    try:
        url = subir_pdf(ruta_temp, carpeta="cotizador/facturas")
        os.remove(ruta_temp)
        return url
    except Exception:
        return nombre_final

@bp.route("/ordenes/archivo/<path:nombre_archivo>")
def ver_archivo_orden(nombre_archivo):
    nombre_archivo = os.path.basename(nombre_archivo)
    return send_from_directory(
        current_app.config["UPLOAD_FOLDER_ORDENES"],
        nombre_archivo
    )
@bp.route("/ordenes/<int:orden_id>/extraer-pdf", methods=["POST"])
def extraer_pdf_orden(orden_id):
    orden = obtener_orden_por_id(orden_id)
    if not orden:
        flash("La orden no existe.", "warning")
        return redirect("/ordenes")

    if not orden.get("archivo_orden"):
        flash("Esta orden no tiene PDF asociado.", "warning")
        return redirect(f"/ordenes/{orden_id}")

    try:
        ruta_pdf = _resolver_ruta_pdf(
            orden["archivo_orden"],
            current_app.config["UPLOAD_FOLDER_ORDENES"],
        )

        resultado = extraer_texto_pdf(ruta_pdf)
        guardar_resultado_extraccion_orden(
            orden_id,
            resultado["texto"],
            resultado["metodo"]
        )

        if resultado["texto"]:
            datos_parseados = parsear_orden_por_plantilla(
                orden.get("plantilla_ocr"),
                resultado["texto"]
            )

            if datos_parseados:
                aplicar_datos_parseados_a_orden(orden_id, datos_parseados)

            if resultado["metodo"] == "pdf":
                flash("Texto extraído directamente del PDF.", "success")
            else:
                flash("Texto extraído con OCR.", "success")

    except Exception as e:
        flash(f"Error al extraer el texto del PDF: {e}", "danger")

    return redirect(f"/ordenes/{orden_id}")    

def guardar_pdf_temporal_orden(archivo):
    if not archivo or not archivo.filename:
        return ""
    if not archivo_permitido_pdf(archivo.filename):
        return ""
    nombre_seguro = secure_filename(archivo.filename)
    nombre_final = f"{uuid4().hex}_{nombre_seguro}"
    ruta_temp = os.path.join(current_app.config["UPLOAD_FOLDER_ORDENES"], nombre_final)
    archivo.save(ruta_temp)
    try:
        url = subir_pdf(ruta_temp, carpeta="cotizador/ordenes")
        os.remove(ruta_temp)  # borra el archivo local después de subir
        return url
    except Exception:
        return nombre_final  # si falla Cloudinary, usa archivo local guardado

@bp.route("/ordenes/nueva/prellenar", methods=["POST"])
def prellenar_nueva_orden():
    empresas = obtener_empresas_activas()

    empresa_id = request.form.get("empresa_id", "").strip()
    numero_orden = request.form.get("numero_orden", "").strip()
    fecha_orden = request.form.get("fecha_orden", "").strip()
    subtotal = request.form.get("subtotal", "").strip()
    iva = request.form.get("iva", "").strip()
    total = request.form.get("total", "").strip()
    moneda = request.form.get("moneda", "GTQ").strip()
    proyecto = request.form.get("proyecto", "").strip()
    descripcion_extraida = request.form.get("descripcion_extraida", "").strip()
    condiciones_pago = request.form.get("condiciones_pago", "").strip()
    dias_credito = request.form.get("dias_credito", "").strip()
    archivo_existente = request.form.get("archivo_orden_guardado", "").strip()

    archivo_orden = archivo_existente
    archivo = request.files.get("archivo_orden")

    if archivo and archivo.filename:
        archivo_orden = guardar_pdf_temporal_orden(archivo)
        if not archivo_orden:
            flash("Solo se permiten archivos PDF para la orden de compra.", "warning")
            return redirect("/ordenes/nueva")

    if not empresa_id:
        flash("Debes seleccionar una empresa antes de extraer datos.", "warning")
        return render_template(
            "orden_nueva.html",
            empresas=empresas,
            form_data=request.form,
            archivo_orden_guardado=archivo_orden,
            datos_detectados={},
        )

    if not archivo_orden:
        flash("Debes subir un PDF para extraer datos.", "warning")
        return render_template(
            "orden_nueva.html",
            empresas=empresas,
            form_data=request.form,
            archivo_orden_guardado="",
            datos_detectados={},
        )

    empresa = obtener_empresa_por_id(int(empresa_id))
    ruta_pdf = _resolver_ruta_pdf(
        archivo_orden,
        current_app.config["UPLOAD_FOLDER_ORDENES"],
    )

    resultado = extraer_texto_pdf(ruta_pdf)
    datos_detectados = {}

    if resultado["texto"]:
        datos_detectados = parsear_orden_por_plantilla(
            empresa.get("plantilla_ocr") if empresa else "",
            resultado["texto"]
        ) or {}

        descripcion_extraida = resultado["texto"]

        if not numero_orden:
            numero_orden = datos_detectados.get("numero_orden") or ""
        if not fecha_orden:
            fecha_orden = datos_detectados.get("fecha_orden") or ""
        if not subtotal:
            subtotal = datos_detectados.get("subtotal") or ""
        if not iva:
            iva = datos_detectados.get("iva") or ""
        if not total:
            total = datos_detectados.get("total") or ""
        if not proyecto:
            proyecto = datos_detectados.get("proyecto") or ""
        if not condiciones_pago:
            condiciones_pago = datos_detectados.get("condiciones_pago") or ""

        flash(
            "Datos extraídos del PDF. Revisa y corrige antes de guardar.",
            "success"
        )
    else:
        flash("No se pudo extraer texto del PDF.", "warning")

    form_data = {
        "empresa_id": empresa_id,
        "nit_cliente": empresa.get("nit", "") if empresa else "",
        "direccion_cliente": empresa.get("direccion", "") if empresa else "",
        "dias_credito": dias_credito or (empresa.get("dias_credito") if empresa else 30),
        "numero_orden": numero_orden,
        "fecha_orden": fecha_orden,
        "subtotal": subtotal,
        "iva": iva,
        "total": total,
        "moneda": moneda,
        "proyecto": proyecto,
        "descripcion_extraida": descripcion_extraida,
        "condiciones_pago": condiciones_pago,
        "descripcion_sat": f"por servicios tecnicos segun orden de compra {numero_orden} {proyecto}".replace("  ", " ").strip(),
    }

    return render_template(
        "orden_nueva.html",
        empresas=empresas,
        form_data=form_data,
        archivo_orden_guardado=archivo_orden,
        datos_detectados=datos_detectados,
    )

@bp.route("/ordenes/<int:orden_id>/factura/prellenar", methods=["POST"])
def prellenar_factura(orden_id):
    orden = obtener_orden_por_id(orden_id)
    if not orden:
        flash("La orden no existe.", "warning")
        return redirect("/ordenes")

    archivo_factura_guardado = request.form.get("archivo_factura_guardado", "").strip()
    archivo = request.files.get("archivo_factura")

    if archivo and archivo.filename:
        archivo_factura_guardado = guardar_pdf_temporal_factura(archivo)
        if not archivo_factura_guardado:
            flash("Solo se permiten archivos PDF para la factura.", "warning")
            return redirect(f"/ordenes/{orden_id}/factura")

    if not archivo_factura_guardado:
        flash("Debes subir el PDF de la factura SAT.", "warning")
        return redirect(f"/ordenes/{orden_id}/factura")

    ruta_pdf = _resolver_ruta_pdf(
        archivo_factura_guardado,
        current_app.config["UPLOAD_FOLDER_FACTURAS"],
    )
    resultado_pdf = extraer_texto_pdf(ruta_pdf)
    datos_detectados = {}

    if resultado_pdf["texto"]:
        datos_detectados = parsear_factura_sat(resultado_pdf["texto"]) or {}
        flash("Datos extraídos de la factura SAT.", "success")
    else:
        flash("No se pudo extraer texto de la factura.", "warning")

    form_data = {
        "numero_factura": datos_detectados.get("numero_factura") or orden.get("numero_factura") or "",
        "fecha_factura": datos_detectados.get("fecha_factura") or orden.get("fecha_factura") or "",
        "monto_facturado": datos_detectados.get("monto_facturado") or orden.get("monto_facturado") or orden.get("total") or "",
        "nit_facturado": orden.get("empresa_nit") or "",
        "direccion_facturada": orden.get("empresa_direccion") or "",
        "descripcion_sat": orden.get("descripcion_sat_sugerida") or "",
        "observaciones": orden.get("observaciones") or "",
    }

    return render_template(
        "registrar_factura.html",
        orden=orden,
        form_data=form_data,
        datos_detectados=datos_detectados,
        archivo_factura_guardado=archivo_factura_guardado,
    )   
@bp.route("/facturas/archivo/<path:nombre_archivo>")
def ver_archivo_factura(nombre_archivo):
    nombre_archivo = os.path.basename(nombre_archivo)
    return send_from_directory(
        current_app.config["UPLOAD_FOLDER_FACTURAS"],
        nombre_archivo
    )     
